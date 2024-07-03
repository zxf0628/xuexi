import datetime
import socket
import openpyxl
import pandas
from PyQt5.QtWidgets import QFileDialog

# 1. 获取基于该文件的上上级目录，为：*\Tcp，os.path.dirname一个嵌套为一级
# 2024.7.3 拿到文件上层目录，设置为环境路径 达到实现导入其他模块
# BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# sys.path.append(BASE_DIR)


class Common_Method():
    def __init__(self):
        pass

    """获取本机IP地址"""

    @staticmethod
    def getHostIP():
        hostip = socket.gethostbyname_ex(socket.gethostname())
        return hostip[-1][-1]

    """将文本文件转为excel"""

    @staticmethod
    def file_transfer_excel(path):
        path = r"{}".format(path)
        print("Common_Method->file_transfer_excel->要转换文件的路径:{}".format(path))

        # 将文本每行插入excel
        work_logs = openpyxl.Workbook()
        sheet = work_logs.create_sheet("logs")
        logdata = open(path, "r", encoding="utf-8")
        logdatas = logdata.readlines()
        for index, row in enumerate(logdatas):
            d = row.split()
            for col in range(len(d)):
                sheet.cell(index + 1, col + 1, d[col])

        Common_Method.derive(work_logs)
        return True

    '''给excel已当前时间点命名，并删除表头的符号，将其导出'''

    @staticmethod
    def derive(work_logs):
        # 导出excel
        curr_datetime = datetime.datetime.now()
        curr_datetime_str = curr_datetime.strftime("%Y,%m,%d,%H,%M")
        file_name = r"./{}.xlsx".format(curr_datetime_str)
        work_logs.save(file_name)

        # 删除数据的开头无用数据  用最新excel覆盖excel
        logsdata = pandas.read_excel(file_name, sheet_name="logs")
        col_name = list(logsdata.columns[1:])
        col_name.append(logsdata.columns[0])
        logsdata.columns = col_name
        logsdata.to_excel(file_name, sheet_name="logs")

    '''打开电脑选择文件，返回文件的目录'''

    @staticmethod
    def getOpenFileName():
        file_name = QFileDialog.getOpenFileName(None, "选择文件路径", "./", "All Files(*)")
        return file_name


    '''将二维元组，转换成为 以,分割的str'''

    @staticmethod
    def typle_transition_to_str(result):
        s1 = ','.join([i[0] for i in result])
        return s1


    '''将二维元组，转换成为 以,分割的list'''

    @staticmethod
    def typle_transition_to_list(result):
        s1 = ','.join([i[0] for i in result])
        return list(s1)


    '''获取当前时间 年月日小时分钟 返回字符串形式'''

    @staticmethod
    def now_time_str():
        curr_datetime = datetime.datetime.now()
        curr_datetime_str = curr_datetime.strftime("%Y,%m,%d,%H,%M")
        return curr_datetime_str
